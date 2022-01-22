import json
import os
import xlrd
import xlwt
import openpyxl
from algorithm import request_extract

def write_result(dir, name, save_name, request_Columns, department_Columns):
    file_path = os.path.join(dir, name)
    save_path = os.path.join(dir, save_name)

    # 打开文件
    workbook = openpyxl.load_workbook(file_path)
    sheetnames = workbook.get_sheet_names()
    sheet = workbook.get_sheet_by_name(sheetnames[0])

    ## 或者如果你只是想创建一张空表
    copy_workbook = openpyxl.Workbook()
    # 创建一个sheet
    copy_sheet = copy_workbook.create_sheet(index=0)

    # 写入一个值，括号内分别为行数、列数、内容
    copy_sheet.cell(1, 1).value = "登记时间"
    copy_sheet.cell(1, 2).value = "诉求"
    copy_sheet.cell(1, 3).value = "目的代码"
    copy_sheet.cell(1, 4).value = "目的"
    copy_sheet.cell(1, 5).value = "分类代码"
    copy_sheet.cell(1, 6).value = "分类"
    copy_sheet.cell(1, 7).value = "预测分类"
    copy_sheet.cell(1, 8).value = "紧急程度"
    copy_sheet.cell(1, 9).value = "危险程度"
    copy_sheet.cell(1, 10).value = "核心词汇"
    copy_sheet.cell(1, 11).value = "诉求问题"
    copy_sheet.cell(1, 12).value = "相关机构"
    copy_sheet.cell(1, 13).value = "关联地址"
    # 遍历
    rows = sheet.max_row
    clos = sheet.max_column
    print('rows:', rows)
    print('clos:', clos)
    for row in range(2, rows + 1):
        copy_sheet.cell(row, 1).value = sheet.cell(row, 1).value
        copy_sheet.cell(row, 2).value = sheet.cell(row, 2).value
        copy_sheet.cell(row, 3).value = sheet.cell(row, 3).value
        copy_sheet.cell(row, 4).value = sheet.cell(row, 4).value
        copy_sheet.cell(row, 5).value = sheet.cell(row, 5).value
        copy_sheet.cell(row, 6).value = sheet.cell(row, 6).value
        newData = {}
        # 获取逐条投诉数据
        content = sheet.cell(row, request_Columns).value
        content = content.replace('\n', '')
        content = content.replace('\r', '')
        content = content.replace('\t', '')
        content = content.replace(' ', '')
        department = sheet.cell(row, department_Columns).value
        # 算法处理
        if request_extract.emergency_degree_classification(content, request_extract.emergency_word) == True:
            newData['degree_of_urgency'] = "紧急"
        else:
            newData['degree_of_urgency'] = "一般"
        newData['issue'] = request_extract.get_request(content, request_extract.request_word,
                                                       request_extract.request_double_word)
        temp_request = request_extract.get_keyinfo(content)
        if newData['issue'] == '':
            newData['issue'] = request_extract.get_request_by_keyword(content)
        newData['keyword'] = temp_request['keywords']
        newData['organization'] = temp_request['org']
        newData['address'] = temp_request['location']
        # [1~5]分别对应['可忽略危险', '临界危险', '一般危险', '破坏性危险', '毁灭性危险']
        degree_of_dangerous = request_extract.dangerous_degree_classification(content, request_extract.dangerous_word)
        if degree_of_dangerous == 1:
            newData['degree_of_dangerous'] = '可忽略危险'
        elif degree_of_dangerous == 2:
            newData['degree_of_dangerous'] = '临界危险'
        elif degree_of_dangerous == 3:
            newData['degree_of_dangerous'] = '一般危险'
        elif degree_of_dangerous == 4:
            newData['degree_of_dangerous'] = '破坏性危险'
        elif degree_of_dangerous == 5:
            newData['degree_of_dangerous'] = '毁灭性危险'

        copy_sheet.cell(row, 8).value = str(newData['degree_of_urgency'])
        copy_sheet.cell(row, 9).value = str(newData['degree_of_dangerous'])
        copy_sheet.cell(row, 10).value = str(newData['keyword'])[1: -1]
        copy_sheet.cell(row, 11).value = str(newData['issue'])
        copy_sheet.cell(row, 12).value = str(newData['organization'])[1: -1]
        copy_sheet.cell(row, 13).value = str(newData['address'])[1: -1]
        print(row)

    copy_workbook.save(save_path)
    print('finish...')

dir = '../static/data/'
name = 'data_excel1.xlsx'
save_name = 'excel1_with_results.xlsx'
request_Columns = 2
department_Columns = 6
write_result(dir, name, save_name, request_Columns, department_Columns)
print('finished..........')